import openpyxl
import sys
import os
import re
import tkinter
from tkinter import filedialog


#あとでGUI入力に帰る
xlsx_path = '2. 2021年5月教室割_0426.xlsx'
myoji = str(sys.argv[1])


#ファイル作成
path = './test.html'
f = open(path, 'a', encoding="utf-8")
f.write("<body bgcolor= black text=white>")

wb = openpyxl.load_workbook(xlsx_path)
ws = wb.worksheets[0]

#表の範囲取得
fmt = "{:7d} {:7d} {:7d} {:7d}"

#行範囲
sta_row = int(ws.min_row)
end_row = int(ws.max_row)

#列範囲
sta_col = int(ws.min_column)
end_col = int(ws.max_column)

#日付書き込み関数
def get_dayinfo(day):
    w_day = ("<font color=aliceblue><h1>" + day + "</h1></font>\n")
    return f.write(w_day)
    f.close()

#時間書き込み関数
def get_time(time):
    w_time = ("<font color=hotpink><br><p1>" + time + "</p1></font>")
    return f.write(w_time)
    f.close

#名前書き込み関数
def get_myname(name):
    w_name = ("<font color=cyan><p1>" + "・・・" + name + "</p1><br></font>")
    return f.write(w_name)
    f.close

#横移動
i = 1#行
n = 1#列

for i in range(1, end_row):#行を1ずつ下がる
#    print("行=" + str(i))
    for n in range(1, end_col):#列を1ずつ右へ
#        print("列=" + str(n))
        celldata = ws.cell(row=i, column=n).value
#        print(celldata)
        if celldata is None:
            pass
        elif '日鏵國際企劃' in celldata:
            pass
        elif '日期' in celldata:
            pass
        elif '時間' in celldata:
            pass
        elif n == 1:#日付を見出しとしてファイルに書き込み
            day = (celldata)
            get_dayinfo(day)
        elif n == 2:
            time = celldata
            get_time(time)
        elif myoji in celldata:
            name = celldata
            get_myname(name)

