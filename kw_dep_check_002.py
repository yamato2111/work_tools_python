import openpyxl

file_path = input('ファイルをドラッグ＆ドロップしてEnter>>')
wb = openpyxl.load_workbook(file_path.replace('\"', ''))
ws = wb.worksheets[0]
print('\n')

#(修正メモ)SQLから教師名読み込み
#とりあえず配列で対応
teachers_name = teachers

#表の範囲取得
fmt = "{:7d} {:7d} {:7d} {:7d}"
max_row = int(ws.max_row)#行MAX
max_col = int(ws.max_column)#列MAX

i = 1#行
n = 1#列

for i in range(1, max_row):#1行ずつ下がる
    name_list = []
    for n in range(1, max_col):#1列ずつ右へ
        celldata = ws.cell(row=i, column=n).value#現在のセル
        if celldata is None or not str(celldata).strip():#セルが空、string型じゃなければスキップ
            continue
        else:
            name_list.append(celldata[-4:])#セルの内容を後ろからスライス
    for teacher in teachers_name:
        if len(name_list) != len(set(name_list)):#重複チェック #メモ：ロジック変更。教師名リストの内容をイテレート。教師名がスライスした内容に含まれていたらprint
            print('・' + str(i) + '行目×\n')

print(name_list)
input('\nEnterキー押下で終了')
