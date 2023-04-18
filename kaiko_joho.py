import openpyxl as ox
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import datetime
import calendar
import re

wb_kai = ox.load_workbook(kaiko_file)
ws_kai = wb_kai.worksheets[0]

#----------開講情報整理--------------------
#「開課日」を探す
row = 1 #行
col = 1 #列
cell_val = ws_kai.cell(row, col).value

while ws_kai.cell(row, col).value != "開課日":
    row += 1
#開始地点へ移動
row += 2
col += 3

kaiko_list = []
i = 0
#
while ws_kai.cell(row, col).value is not None:
    k_list_name = "list" + str(i)
    k_list_name = list()#リスト化
    kurasu_date = ws_kai.cell(row, col).value#開始地点
    k_list_name.append(kurasu_date)
    col -= 3
    kurasu_name = ws_kai.cell(row, col).value
    key = "]"
    if "綜合" in kurasu_name:
        col += 1
        sogo_level = ws_kai.cell(row, col).value
        kakko = kurasu_name.find(key)
        k_Id = kurasu_name[1:kakko]
        aft_kakko = kakko + 2
        k_Name = kurasu_name[aft_kakko:]
        k_list_name.append(k_Id)
        k_list_name.append(k_Name)
        k_list_name.append(sogo_level)
        col -= 1
    elif "個人班" in kurasu_name:
        col += 1
        kojin_level = ws_kai.cell(row, col).value
        kakko = kurasu_name.find(key)
        k_Id = kurasu_name[1:kakko]
        aft_kakko = kakko + 2
        k_Name = kurasu_name[aft_kakko:]
        k_list_name.append(k_Id)
        k_list_name.append(k_Name)
        k_list_name.append(kojin_level)
        col -= 1
    elif "同步課程" in kurasu_name:
        col += 1
        doho_level = ws_kai.cell(row, col).value
        kakko = kurasu_name.find(key)
        k_Id = kurasu_name[1:kakko]
        aft_kakko = kakko + 2
        k_Name = kurasu_name[aft_kakko:]
        k_list_name.append(k_Id)
        k_list_name.append(k_Name)
        k_list_name.append(doho_level)
        col -= 1
    else:
        kakko = kurasu_name.find(key)
        k_Id = kurasu_name[1:kakko]
        aft_kakko = kakko + 2
        k_Name = kurasu_name[aft_kakko:]
        k_list_name.append(k_Id)
        k_list_name.append(k_Name)
    col += 3
    row += 1
    kaiko_list.append(k_list_name)
#----------教室割ファイル検索--------------------
wb_kw = ox.load_workbook(schedule_file)
ws_kw = wb_kw.worksheets[0]
#行範囲
end_row = int(ws_kw.max_row)

#列範囲
end_col = int(ws_kw.max_column)

#横移動
r = 1#行
c = 1#列
celldata = ws_kw.cell(row=r, column=c).value
#二次元リスト、内包リストの要素番号
y = 0
ite_kaiko_list = iter(kaiko_list)

for key_list in ite_kaiko_list: #探索用の授業リスト取り出し
    day_break = False
    brk = False
    key_day = key_list[0] #開講日
    key_Id = key_list[1] #授業ID
    r = 3
    for r in range(r, end_row): #授業日の列だけ探索
        if day_break == True:
            break
        day_cell = ws_kw.cell(row=r, column=1).value
        if day_cell is None:
            continue
        elif key_day in day_cell:
            for r in range(r, end_row): #行を1ずつ下がる
                if brk == True:
                    break
                for c in range(1, end_col): #列を1ずつ右へ
                    celldata = ws_kw.cell(row=r, column=c).value
                    if celldata is None:
                        continue
                    elif key_Id in celldata:
                        t_name = celldata[-4:]
                        key_list.append(t_name)
                        day_break = True
                        brk = True
                        break

#----------案内情報整理--------------------
wb_ann = ox.load_workbook(kaiko_file)
ws_ann = wb_kai.worksheets[0]

#「案内日」を探す
row = 1 #行
col = 1 #列
cell_val = ws_ann.cell(row, col).value

while ws_ann.cell(row, col).value != "案內日":
    row += 1
#開始地点へ移動
row += 2
col += 4

annai_list = []
i = 0
#
while ws_ann.cell(row, col).value is not None:
    a_list_name = "list" + str(i)
    a_list_name = list()#リスト化
    kurasu_date = ws_ann.cell(row, col).value#開始地点
    a_list_name.append(kurasu_date)
    col -= 4
    kurasu_name = ws_ann.cell(row, col).value
    key = "]"
    if "綜合" in kurasu_name:
        col += 1
        sogo_level = ws_ann.cell(row, col).value
        kakko = kurasu_name.find(key)
        a_Id = kurasu_name[1:kakko]
        aft_kakko = kakko + 2
        a_Name = kurasu_name[aft_kakko:]
        a_list_name.append(a_Id)
        a_list_name.append(a_Name)
        a_list_name.append(sogo_level)
        col -= 1
    elif "個人班" in kurasu_name:
        col += 1
        kojin_level = ws_ann.cell(row, col).value
        kakko = kurasu_name.find(key)
        a_Id = kurasu_name[1:kakko]
        aft_kakko = kakko + 2
        a_Name = kurasu_name[aft_kakko:]
        a_list_name.append(a_Id)
        a_list_name.append(a_Name)
        a_list_name.append(kojin_level)
        col -= 1
    elif "同步課程" in kurasu_name:
        col += 1
        doho_level = ws_ann.cell(row, col).value
        kakko = kurasu_name.find(key)
        a_Id = kurasu_name[1:kakko]
        aft_kakko = kakko + 2
        a_Name = kurasu_name[aft_kakko:]
        a_list_name.append(a_Id)
        a_list_name.append(a_Name)
        a_list_name.append(doho_level)
        col -= 1
    else:
        kakko = kurasu_name.find(key)
        a_Id = kurasu_name[1:kakko]
        aft_kakko = kakko + 2
        a_Name = kurasu_name[aft_kakko:]
        a_list_name.append(a_Id)
        a_list_name.append(a_Name)
    col += 4
    row += 1
    annai_list.append(a_list_name)
#----------教室割ファイル検索--------------------
wb_kw = ox.load_workbook(schedule_file)
ws_kw = wb_kw.worksheets[0]

#行範囲
end_row = int(ws_kw.max_row)

#列範囲
end_col = int(ws_kw.max_column)

#横移動
r = 1#行
c = 1#列
celldata = ws_kw.cell(row=r, column=c).value
#二次元リスト、内包リストの要素番号
y = 0
ite_annai_list = iter(annai_list)

for key_list in ite_annai_list: #探索用の授業リスト取り出し
    day_break = False
    brk = False
    key_day = key_list[0] #開講日
    key_Id = key_list[1] #授業ID
    r = 1
    for r in range(r, end_row): #授業日の列だけ探索
        if day_break == True:
            break
        day_cell = ws_kw.cell(row=r, column=1).value
        if day_cell is None:
            continue
        elif key_day in day_cell:
            for r in range(r, end_row): #行を1ずつ下がる
                if brk == True:
                    break
                for c in range(1, end_col): #列を1ずつ右へ
                    celldata = ws_kw.cell(row=r, column=c).value
                    if celldata is None:
                        continue
                    elif key_Id in celldata:
                        a_name = celldata[-4:]
                        key_list.append(a_name)
                        day_break = True
                        brk = True
                        break

#----------説明会検索--------------------
wb_st = ox.load_workbook(schedule_file)
ws_st = wb_st.worksheets[0]

#正規表現
ptr_day = "[0-9][0-9]/[0-9][0-9]\(.\)"
ptr_hui = "[0-9][0-9]:[0-9][0-9] ~ [0-9][0-9]:[0-9][0-9]"
pattern_day = re.compile(ptr_day)
pattern_hui = re.compile(ptr_hui)

#行範囲
end_row = int(ws_st.max_row)
#列範囲
end_col = 100

#大枠リスト
setsu_list = []
i = 0 #内包リスト追加用

r = 1#行
c = 1#列
celldata = ws_st.cell(row=r, column=c).value
r = 1
c = 1
for r in range(r, end_row): #行を1ずつ下がる
    c = 1
    for c in range(c, end_col): #列を1ずつ右へ
        celldata = ws_st.cell(row=r, column=c).value # セルの現在地
        if celldata is None:
            continue
        else:
            res = pattern_hui.search(celldata)
            if res is not None:
                s_list_name = "list" + str(i)
                s_list_name = list()#リスト化
                rr = r
                cell_d = ws_st.cell(row=rr, column=1).value #説明会の日付を探索
                while cell_d is None:
                    rr -= 1
                    cell_d = ws_st.cell(row=rr, column=1).value
                s_list_name.append(cell_d[0:5])
                s_list_name.append(celldata)
                setsu_list.append(s_list_name)

#-----------------------------------------------------------表作成
title_alignment=Alignment(horizontal='left', # セルサイズ
                     vertical='center',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)

cell_alignment=Alignment(horizontal='left', # セルサイズ
                     vertical='top',
                     text_rotation=0,
                     wrap_text=True,
                     shrink_to_fit=False,
                     indent=0)

#今月（入力仕様後で作る）
this_year = 2021
this_month = 5
#月の最終日取得
end_date = calendar.monthrange(this_year, this_month)[1]
# ブック新規作成
wb_hyo = ox.Workbook()
# シート指定
ws_hyo = wb_hyo.worksheets[0]
# 罫線(外枠)を設定
border = Border(top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'))
# セルに罫線を設定
dd = end_date + 3 #日付分の行数（ループ用）
for row_num in range(2,dd):
    for col_num in range(2,6):
        ws_hyo.cell(row=row_num ,column=col_num).border = border

font = Font(size=14, bold=True)

# 高さを設定
ws_hyo.row_dimensions[2].height = 30
# 幅を設定
ws_hyo.column_dimensions['B'].width = 10
ws_hyo.column_dimensions['C'].width = 40
ws_hyo.column_dimensions['D'].width = 40
ws_hyo.column_dimensions['E'].width = 40


#見出し書き込み
ws_hyo["B2"].value = "日付"
ws_hyo["B2"].alignment = title_alignment
ws_hyo["B2"].font = font
ws_hyo["C2"].value = "開講"
ws_hyo["C2"].alignment = title_alignment
ws_hyo["C2"].font = font
ws_hyo["D2"].value = "案内"
ws_hyo["D2"].alignment = title_alignment
ws_hyo["D2"].font = font
ws_hyo["E2"].value = "説明会"
ws_hyo["E2"].alignment = title_alignment
ws_hyo["E2"].font = font

#日付書き込み
this_month = str(this_month)

hyo_date_list = []
for date in range(1, end_date+1):
    date = str(date)
    date.zfill(2)
    mon_d = "{0}/{1}".format(this_month.zfill(2), date.zfill(2))
    hyo_date_list.append(mon_d)
dr = 3
for h in range(len(hyo_date_list)):
    ws_hyo.cell(row=dr, column=2).value = hyo_date_list[h]
    ws_hyo.cell(row=dr, column=2).alignment = cell_alignment
    dr += 1

#　開講
end_row = int(ws_hyo.max_row)
a = 0
b = 0
for i in range(len(kaiko_list)):
    kai_day = kaiko_list[a][b]
    for ii in range(end_row):
        ii += 1
        celldata = ws_hyo.cell(row=ii, column=2).value
        if celldata == kai_day:
            if len(kaiko_list[a]) == 4:
                org_cell = ws_hyo.cell(row=ii, column=3).value
                add_cell = kaiko_list[a][2] + '/' + kaiko_list[a][3]
                mix_cell = str(org_cell) + str(add_cell)
                ws_hyo.cell(row=ii, column=3).value = mix_cell
            elif len(kaiko_list[a]) == 5:
                org_cell = ws_hyo.cell(row=ii, column=3).value
                add_cell = kaiko_list[a][2] + '・' + kaiko_list[a][3] + '/' + kaiko_list[a][4]
                mix_cell = str(org_cell) + str(add_cell)
                ws_hyo.cell(row=ii, column=3).value = mix_cell
            ws_hyo.cell(row=ii, column=3).alignment = cell_alignment
    a += 1

#　案内
end_row = int(ws_hyo.max_row)
a = 0
b = 0
for i in range(len(annai_list)):
    ann_day = annai_list[a][b]
    for ii in range(end_row):
        ii += 1
        celldata = ws_hyo.cell(row=ii, column=2).value
        if celldata == ann_day:
            if len(annai_list[a]) == 4:
                org_cell = ws_hyo.cell(row=ii, column=4).value
                add_cell = annai_list[a][2] + '/' + annai_list[a][3]
                mix_cell = str(org_cell) + str(add_cell)
                ws_hyo.cell(row=ii, column=4).value = mix_cell
            elif len(annai_list[a]) == 5:
                org_cell = ws_hyo.cell(row=ii, column=4).value
                add_cell = annai_list[a][2] + '・' + annai_list[a][3] + '/' + annai_list[a][4]
                mix_cell = str(org_cell) + str(add_cell)
                ws_hyo.cell(row=ii, column=4).value = mix_cell
            ws_hyo.cell(row=ii, column=4).alignment = cell_alignment
    a += 1

#　説明会
end_row = int(ws_hyo.max_row)
a = 0
b = 0
for i in range(len(setsu_list)):
    sets_day = setsu_list[a][b]
    for ii in range(end_row):
        ii += 1
        celldata = ws_hyo.cell(row=ii, column=2).value
        if celldata == sets_day:
            org_cell = ws_hyo.cell(row=ii, column=5).value
            add_cell = setsu_list[a][1]
            mix_cell = str(org_cell) + '・' + str(add_cell)
            ws_hyo.cell(row=ii, column=5).value = mix_cell
        ws_hyo.cell(row=ii, column=5).alignment = cell_alignment
    a += 1

#表を整理
#正規表現
r = 3
end_col_h = 6
for r in range(3, end_row+1): #行を1ずつ下がる
    c = 3
    for c in range(3, end_col_h): #列を1ずつ右へ
        celldata = ws_hyo.cell(row=r, column=c).value
        if celldata is None:
            ws_hyo.cell(row=r, column=c).value = "なし"
        elif "None" in celldata:
            del_None = re.sub("None", "", celldata)
            ws_hyo.cell(row=r, column=c).value = del_None
        ws_hyo.cell(row=r, column=c).alignment = cell_alignment
wb_hyo.save("開講情報.xlsx")
